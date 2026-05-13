from __future__ import annotations

import csv
import re
from pathlib import Path

from journal_agent.data.repository import JournalRepository
from journal_agent.ingestion.manuscript_parser import ManuscriptParser
from journal_agent.models.schemas import ManuscriptProfile, RecommendationResult
from journal_agent.ranking.scoring import CorpusScoringEngine
from journal_agent.ranking.supervised import SupervisedJournalRanker
from journal_agent.utils.text_processing import detect_language, normalize_space, normalize_title_key


DIRECT_LAW_JOURNAL_TERMS = {
    "law",
    "legal",
    "jurisprudence",
    "court",
    "judicial",
    "justice",
    "criminology",
    "penology",
    "human rights",
    "constitutional",
    "public law",
}
LAW_ADJACENT_JOURNAL_TERMS = {
    "political science",
    "public administration",
    "international relations",
    "governance",
    "regulation",
    "policy",
    "state",
    "institutions",
}


class JournalRecommendationAgent:
    def __init__(self) -> None:
        self.repository = JournalRepository()
        self.parser = ManuscriptParser()

    def recommend(
        self,
        *,
        dataset_path: str | Path,
        taxonomy_path: str | Path,
        model_path: str | Path | None = None,
        manuscript_path: str | Path | None = None,
        title: str | None = None,
        abstract: str | None = None,
        keywords: list[str] | str | None = None,
        discipline: str = "law",
        top_k: int = 15,
        candidate_scope: str = "law-related",
        scopus_source_list: str | Path | None = None,
        wos_source_list: str | Path | None = None,
    ) -> tuple[ManuscriptProfile, list[RecommendationResult]]:
        manuscript = self.parser.parse(
            manuscript_path,
            title=title,
            abstract=abstract,
            keywords=keywords,
            discipline=discipline,
        )
        journals = self.repository.load_journals(dataset_path, discipline=discipline)
        taxonomy = self.repository.load_taxonomy(taxonomy_path)
        engine = CorpusScoringEngine(taxonomy)
        manuscript = engine.enrich_manuscript(manuscript)
        if model_path:
            ranker = SupervisedJournalRanker.load(model_path)
            journals = self._select_candidate_journals(
                journals,
                manuscript,
                discipline=discipline,
                candidate_scope=candidate_scope,
                scopus_source_list=scopus_source_list,
                wos_source_list=wos_source_list,
            )
            if not journals:
                raise ValueError(
                    "No candidate journals remain after language filtering. "
                    "Check that the dataset contains journals in the manuscript language."
                )
            recommendations = ranker.recommend(manuscript, candidate_journals=journals, top_k=top_k)
            return manuscript, recommendations[:top_k]

        journals = self._select_candidate_journals(
            journals,
            manuscript,
            discipline=discipline,
            candidate_scope=candidate_scope,
            scopus_source_list=scopus_source_list,
            wos_source_list=wos_source_list,
        )
        if not journals:
            raise ValueError(
                "No candidate journals remain after language filtering. "
                "Check that the dataset contains journals in the manuscript language."
            )
        recommendations = engine.score(manuscript, journals)
        return manuscript, recommendations[:top_k]

    def export_results(self, recommendations: list[RecommendationResult], output_path: str | Path) -> None:
        path = Path(output_path)
        if path.suffix.lower() == ".xlsx":
            raise ValueError("Only CSV export is supported. Please use a .csv output path.")
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = [item.as_csv_row() for item in recommendations]
        if not rows:
            raise ValueError("No recommendations to export.")
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    def _select_candidate_journals(
        self,
        journals: list,
        manuscript: ManuscriptProfile,
        *,
        discipline: str,
        candidate_scope: str = "law-related",
        scopus_source_list: str | Path | None = None,
        wos_source_list: str | Path | None = None,
    ) -> list:
        manuscript_language = manuscript.language
        broad_candidates = []
        narrow_candidates = []
        legal_focus = manuscript.legal_topic_score >= 0.55
        scopus_law_ids = (
            self._scopus_law_journal_ids(journals, scopus_source_list)
            if candidate_scope == "scopus-law"
            else set()
        )
        wos_law_ids = (
            self._wos_law_journal_ids(journals, wos_source_list)
            if candidate_scope == "wos-law"
            else set()
        )
        for journal in journals:
            journal_language = self._journal_language(journal)
            if discipline == "law":
                journal_id = journal.journal_id or normalize_title_key(journal.title)
                if candidate_scope == "scopus-law":
                    if journal_id not in scopus_law_ids:
                        continue
                elif candidate_scope == "wos-law":
                    if journal_id not in wos_law_ids:
                        continue
                elif candidate_scope == "law-only":
                    if not self._is_direct_law_journal(journal):
                        continue
                elif not self._is_law_related_journal(journal):
                    continue
                if manuscript_language in {"zh", "en"} and journal_language != manuscript_language:
                    continue
                broad_candidates.append(journal)
                narrow_candidates.append(journal)
                continue
            if manuscript_language == "zh":
                if journal_language != "zh":
                    continue
                broad_candidates.append(journal)
                if self._is_law_related_journal(journal):
                    narrow_candidates.append(journal)
                continue
            if manuscript_language == "en":
                if journal_language != "en":
                    continue
                if self._is_humanities_social_sciences(journal) or journal.discipline == discipline or self._is_ssci(journal):
                    broad_candidates.append(journal)
                if self._is_law_related_journal(journal):
                    narrow_candidates.append(journal)
                continue
            broad_candidates.append(journal)
            if self._is_law_related_journal(journal):
                narrow_candidates.append(journal)
        if legal_focus and narrow_candidates:
            return narrow_candidates
        return broad_candidates

    def _journal_language(self, journal) -> str:
        if journal.language:
            normalized = journal.language.strip().lower()
            if normalized.startswith("zh") or "chinese" in normalized:
                return "zh"
            if normalized.startswith("en") or "english" in normalized:
                return "en"
        return detect_language("\n".join([journal.title, journal.aims_and_scope]))

    def _is_ssci(self, journal) -> bool:
        return any(index.strip().upper() == "SSCI" for index in journal.indexing)

    def _is_humanities_social_sciences(self, journal) -> bool:
        discipline = (journal.discipline or "").strip().lower()
        if discipline in {
            "law",
            "political science",
            "public policy",
            "sociology",
            "anthropology",
            "economics",
            "history",
            "philosophy",
            "international relations",
            "area studies",
            "criminology",
            "communication",
            "education",
            "social sciences",
            "humanities",
        }:
            return True
        index_set = {item.strip().upper() for item in journal.indexing}
        return bool(index_set.intersection({"SSCI", "AHCI", "ESCI"}))

    def _is_law_related_journal(self, journal) -> bool:
        journal_text = " ".join(
            [
                journal.title,
                journal.discipline or "",
                " ".join(journal.subdisciplines),
                " ".join(journal.keywords),
                journal.aims_and_scope,
            ]
        ).lower()
        if any(term in journal_text for term in DIRECT_LAW_JOURNAL_TERMS):
            return True
        if any(term in journal_text for term in LAW_ADJACENT_JOURNAL_TERMS):
            return True
        return False

    def _is_direct_law_journal(self, journal) -> bool:
        journal_text = " ".join(
            [
                journal.title,
                " ".join(journal.subdisciplines),
                " ".join(journal.keywords),
            ]
        ).lower()
        return any(term in journal_text for term in DIRECT_LAW_JOURNAL_TERMS)

    def _scopus_law_journal_ids(self, journals: list, source_list: str | Path | None) -> set[str]:
        if source_list is None:
            raise ValueError("Use --scopus-source-list when --candidate-scope scopus-law is selected.")
        path = Path(source_list)
        if not path.exists():
            raise ValueError(
                f"Scopus source list not found: {path}. "
                "Download the Scopus Source List XLSX and pass it with --scopus-source-list."
            )
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Install openpyxl to use --candidate-scope scopus-law.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["Scopus Sources Mar. 2026"] if "Scopus Sources Mar. 2026" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]

        def column_index(name: str) -> int:
            try:
                return headers.index(name)
            except ValueError as exc:
                raise ValueError(f"Scopus source list is missing expected column: {name}") from exc

        title_col = column_index("Source Title")
        issn_col = column_index("ISSN")
        eissn_col = column_index("EISSN")
        status_col = column_index("Active or Inactive")
        type_col = column_index("Source Type")
        asjc_col = column_index("All Science Journal Classification Codes (ASJC)")

        scopus_titles: set[str] = set()
        scopus_issns: set[str] = set()
        for row in rows:
            source_type = normalize_space(str(row[type_col] or "")).lower()
            status = normalize_space(str(row[status_col] or "")).lower()
            asjc_codes = set(re.findall(r"\d{4}", normalize_space(str(row[asjc_col] or ""))))
            if source_type != "journal" or status != "active" or "3308" not in asjc_codes:
                continue
            title = normalize_space(str(row[title_col] or ""))
            if title:
                scopus_titles.add(normalize_title_key(title))
            scopus_issns.update(self._split_issns(row[issn_col] if issn_col < len(row) else ""))
            scopus_issns.update(self._split_issns(row[eissn_col] if eissn_col < len(row) else ""))

        selected_ids: set[str] = set()
        for journal in journals:
            journal_id = journal.journal_id or normalize_title_key(journal.title)
            journal_issns = {
                self._normalize_issn(journal.issn),
                self._normalize_issn(journal.eissn),
            }
            if normalize_title_key(journal.title) in scopus_titles or bool(journal_issns & scopus_issns):
                selected_ids.add(journal_id)
        if not selected_ids:
            raise ValueError("No dataset journals matched active Scopus ASJC Law records.")
        return selected_ids

    def _wos_law_journal_ids(self, journals: list, source_list: str | Path | None) -> set[str]:
        if source_list is None:
            raise ValueError("Use --wos-ssci-list when --candidate-scope wos-law is selected.")
        path = Path(source_list)
        if not path.exists():
            raise ValueError(
                f"Web of Science SSCI list not found: {path}. "
                "Pass the local SSCI CSV with --wos-ssci-list."
            )

        wos_titles: set[str] = set()
        wos_issns: set[str] = set()
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                categories = self._split_wos_categories(row.get("Web of Science Categories", ""))
                if "law" not in categories:
                    continue
                title = normalize_space(row.get("Journal title", ""))
                if title:
                    wos_titles.add(normalize_title_key(title))
                wos_issns.update(self._split_issns(row.get("ISSN", "")))
                wos_issns.update(self._split_issns(row.get("eISSN", "")))

        selected_ids: set[str] = set()
        for journal in journals:
            journal_id = journal.journal_id or normalize_title_key(journal.title)
            journal_issns = {
                self._normalize_issn(journal.issn),
                self._normalize_issn(journal.eissn),
            }
            if normalize_title_key(journal.title) in wos_titles or bool(journal_issns & wos_issns):
                selected_ids.add(journal_id)
        if not selected_ids:
            raise ValueError("No dataset journals matched Web of Science SSCI Law records.")
        return selected_ids

    def _split_wos_categories(self, value: object) -> set[str]:
        text = normalize_space(str(value or ""))
        if not text or text.lower() == "nan":
            return set()
        return {item.strip().lower() for item in re.split(r"\s*\|\s*|;", text) if item.strip()}

    def _split_issns(self, value: object) -> set[str]:
        text = normalize_space(str(value or ""))
        if not text or text.lower() == "nan":
            return set()
        return {normalized for part in re.split(r"[,;/\s]+", text) if (normalized := self._normalize_issn(part))}

    def _normalize_issn(self, value: object) -> str:
        text = normalize_space(str(value or ""))
        if not text or text.lower() == "nan":
            return ""
        return re.sub(r"[^0-9Xx]", "", text).upper()
